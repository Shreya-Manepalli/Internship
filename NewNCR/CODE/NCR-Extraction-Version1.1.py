#!/usr/bin/env python



'''
Description: This code extract all the neccessary values from the Non Conformance Reports and exports it into an excel file. 
It accepts a folder or a file as an input.It segregates the files in a folder based on the Drwaing number and outputs the information 
in each file into an excel sheet in the desired format.

'''


__author__ = "Shreya Manepalli"
__version__ = "1.1.0"
__email__ = "Shreya.Manepalli@gknaerospace.com"



############################################################# CODE #############################################################

#IMPORTING LIBRARIES
import tabula
import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook
import os
import glob
import PyPDF2
import re



#Function to iterate across all the files in the folder
def folder_list(folder_path):
    files_list = []

    pdf_files = glob.glob(f"{folder_path}/*.pdf")
    if pdf_files:
        for file_path in pdf_files:
            data = file_path
            files_list.append(data)

    else:
        print("No PDF files found in the folder.")

    return files_list



#Function to check if the sheet name in excel is valid or not
def valid_sheet(string):
    characters_to_remove = ['*', ',', '.', '?', '/', '\\', ':', '[', ']']
    for char in characters_to_remove:
        string = string.replace(char, ' ')
    return string



#Function to check if input is null
def not_null_input(input_value):
    if input_value is None:
        return "None"
    else:
        return str(input_value)
    


#Function to read tables in pdf
def read_pdf(pdf_path, excel_path):
    dfs = tabula.read_pdf(pdf_path, pages='all')
    with pd.ExcelWriter(excel_path) as writer:
        for i, df in enumerate(dfs):
            df.to_excel(writer, sheet_name=f"Sheet{i+1}", index=False)

    workbook = load_workbook(excel_path)
    return workbook



#Function to read tables in first page in the pdf
def read_pdf_page1_table1(workbook):

    workbook 
    sheet_names = workbook.sheetnames

    # NCR, Material, Design (in order)
    cells_table1 = [[2,4], [4,1], [6,3]]
    total_cells_table1 = len(cells_table1)

    table1 = []
    j = 0 
        
    sheet = workbook[sheet_names[0]]

        
    for j in range(total_cells_table1):
        row_number = cells_table1[j][0]
        column_number = cells_table1[j][1]
        cell_value = sheet.cell(row=row_number, column=column_number).value
        data = not_null_input(cell_value)
        table1.append(data)

    return table1



#Function to append the data of the files in the excel sheet by finding the last filled row 
def find_last_row(file_path, sheet_name, column_index):
    try:
        workbook = load_workbook(file_path)
    except FileNotFoundError:
        return 0
    
    # Check if the sheet exists
    if sheet_name not in workbook.sheetnames:
        return 0

    sheet = workbook[sheet_name]
    max_row = sheet.max_row
    
    for row in range(max_row, 0, -1):
        cell_value = sheet.cell(row=row, column=column_index).value
        if cell_value is not None:
            return row

    return 0  # If no filled rows are found



#Function to append the data of the files in the excel sheet by finding the last filled row
def find_last_row2(workbook, column_index, sheet_name):
    
    # Check if the sheet exists
    if sheet_name not in workbook.sheetnames:
        print("sheet name not found in workbook")
        return 0

    sheet = workbook[sheet_name]
    max_row = sheet.max_row
    
    for row in range(max_row, 0, -1):
        cell_value = sheet.cell(row=row, column=column_index).value
        if cell_value is not None:
            return row

    print("Cell value is null")
    return 0  # If no filled rows are found



#Function to load tables in first page of the pdf to excel
def read_pdf_page1_table2(workbook):

    sheet_names = workbook.sheetnames
    column_number = 1
    sheet = workbook[sheet_names[1]]

    table2 = []
    j = 2   
    cell_value = ""  

    none_count = 0

    while none_count < 4:
        row_number = j
        column_number = 1
        cell_value = sheet.cell(row=row_number, column=column_number).value
        data = not_null_input(cell_value)
        if cell_value != None:
            table2.append(data)
        else:
            none_count = none_count + 1
        j = j + 1
    
    return table2



#Function to extract Description of Non Conformance,Charact No. , Defect Type 
def read_pdf_tables_3_4_5(pdf_file):
    text_between_phrases1 = []
    text_between_phrases2 = []
    text_between_phrases3 = []
    
    # Open the PDF file
    with open(pdf_file, 'rb') as file:
        pdf_reader = PyPDF2.PdfReader(file)
        
        # Iterate over each page in the PDF
        for page_num in range(len(pdf_reader.pages)):
            page = pdf_reader.pages[page_num]
            page_text = page.extract_text()
            
            # Find matches between the phrases using regex
            matches1 = re.findall(r"Description of Nonconformance(.*?)Cause of Nonconformance", page_text, re.DOTALL)
            matches2 = re.findall(r"Charact.No(.*?)Requirement Location", page_text, re.DOTALL)
            matches3 = re.findall(r"Defect type(.*?)Charact.No", page_text, re.DOTALL)
            
            # Add the matches to the result list
            text_between_phrases1.extend(matches1)
            text_between_phrases2.extend(matches2)
            text_between_phrases3.extend(matches3)
    
    return text_between_phrases1, text_between_phrases2, text_between_phrases3




#Function to write the extracted values into an excel
def write_to_excel(string, row_no, col_no, file_path, sheet_name):
    try:
        workbook = load_workbook(file_path)
    except FileNotFoundError:
        workbook = Workbook()

    if sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
    else:
        sheet = workbook.create_sheet(title=sheet_name)

    sheet.cell(row=row_no, column=col_no).value = string
    workbook.save(file_path)




#Function to write the extracted values into an excel
def setup_output_excel(sheet_name, output_path):

    output_sheet_name = valid_sheet(str(sheet_name))
    column_index = 1

    start_point = 0

    last_filled_row = find_last_row(output_path, output_sheet_name, column_index)

    if last_filled_row == 0:
        start_point = last_filled_row 
    else:
        start_point = last_filled_row + 1

    if start_point == 0:
        write_to_excel("Item No.", 1 + start_point, 1, output_path, output_sheet_name)
        write_to_excel("Material No.", 1 + start_point, 2, output_path, output_sheet_name)
        write_to_excel("Serial numbers affected", 1 + start_point, 3, output_path, output_sheet_name)
        write_to_excel("NCR-No.", 1 + start_point, 4, output_path, output_sheet_name)
        write_to_excel("Charact. No.", 1 + start_point, 5, output_path, output_sheet_name)
        write_to_excel("Defect Type", 1 + start_point, 6, output_path, output_sheet_name)
        write_to_excel("Description of Nonconformance", 1 + start_point, 7, output_path, output_sheet_name)
        write_to_excel("Measured Value", 1 + start_point, 8, output_path, output_sheet_name)
        write_to_excel("Deviation", 1 + start_point, 9, output_path, output_sheet_name)
        return start_point + 2
    else:
        return start_point + 1




#Function to print NCR No. , Material No.
def print_group_1(table1, start, count, output_path, sheet_name):
    ncr_no = table1[0]
    material_no = table1[1]

    sheet_name = valid_sheet(sheet_name)
    i = 0
    for i in range(count):
        input_str1 = str(ncr_no)
        input_str2 = str(material_no)
        write_to_excel(input_str1, start, 4, output_path, sheet_name)
        write_to_excel(input_str2, start, 2, output_path, sheet_name)
        start = start + 1



#Function to print Item No.
def print_group_2(table2, start, count, output_path, sheet_name):

    if count != len(table2):
        print("Invalid Input for Item No.")
    else:
        sheet_name = valid_sheet(sheet_name)
        i = 0
        for i in range(count):
            input_str = str(table2[i])
            write_to_excel(input_str, start, 1, output_path, sheet_name)
            start = start + 1



#Function to print Description of Non Conformance, Charact No. , Defect Type       
def print_group_3_4_5(table3, table4, table5, start_point, count, output_path, sheet_name):

    start = start_point 

    if count != len(table3):
        print("Invalid Input for Description")
    else:
        start = start_point 
        sheet_name = valid_sheet(sheet_name)
        i = 0
        for i in range(count):
            input_str = str(table3[i])
            write_to_excel(input_str, start, 7, output_path, sheet_name)
            start = start + 1
    
    if count != len(table4):
        print("Invalid Input for Character No.")
    else:
        start = start_point 
        sheet_name = valid_sheet(sheet_name)
        i = 0
        for i in range(count):
            input_str = str(table4[i])
            write_to_excel(input_str, start, 5, output_path, sheet_name)
            start = start + 1
    
    if count != len(table5):
        print("Invalid Input for Defect Type")
    else:
        start = start_point 
        sheet_name = valid_sheet(sheet_name)
        i = 0
        for i in range(count):
            input_str = str(table5[i])
            write_to_excel(input_str, start, 6, output_path, sheet_name)
            start = start + 1



#Function to print Measured Value, Deviation Value
def print_group_6_7(table6, table7, start_point, count, output_path, sheet_name):

    start = start_point 

    if count != len(table6):
        print("Invalid Input for Measured Value")
    else:
        start = start_point 
        sheet_name = valid_sheet(sheet_name)
        i = 0
        for i in range(count):
            input_str = str(table6[i])
            write_to_excel(input_str, start, 8, output_path, sheet_name)
            start = start + 1
    
    if count != len(table7):
        print("Invalid Input for Deviation")
    else:
        start = start_point 
        sheet_name = valid_sheet(sheet_name)
        i = 0
        for i in range(count):
            input_str = str(table7[i])
            write_to_excel(input_str, start, 9, output_path, sheet_name)
            start = start + 1


#Function to extract the measured and deviation values
def get_dev_and_mes(sentences):
    pattern = r"is\s+([\s\S]*?)\s+or\s+(\d+\.?\d*)"

    measured_values = []
    deviation_values = []

    for sentence in sentences:
        if "casting" in sentence.lower():
            measured_value = ""
            deviation_value = ""
        elif sentence.startswith("All"):
            # Search for the pattern in the sentence
            pattern1 = r"at\s+([\s\S]*?)\s+or\s+(\d+\.?\d*)"
            match1 = re.search(pattern1, sentence)

            if match1:
                measured_value = match1.group(1)
                deviation_value = match1.group(2)
            else:
                measured_value = ""
                deviation_value = ""
        else:
            # Search for the pattern in the sentence
            match = re.search(pattern, sentence)

            if match:
                measured_value = match.group(1)
                deviation_value = match.group(2)
            else:
                measured_value = ""
                deviation_value = ""

        measured_values.append(measured_value)
        deviation_values.append(deviation_value)

    table6 = measured_values
    table7 = deviation_values

    return table6, table7



#Function to perform all the file actions
def perform_all_file_actions(pdf_path, output_path):

    excel_path = "./temp_out_f16f157a42d34514f4e545e747b6e880e402cb813f65c81c8dfae2a3e47efdb3.xlsx"

    workbook = read_pdf(pdf_path, excel_path)
    table1 = read_pdf_page1_table1(workbook)
    table2 = read_pdf_page1_table2(workbook)
    sheet_name = table1[2]
    start = setup_output_excel(str(sheet_name), output_path)
    count = len(table2)


    print_group_2(table2, start, count, output_path, sheet_name)
    print_group_1(table1, start, count, output_path, sheet_name)

    table3, table4, table5 = read_pdf_tables_3_4_5(pdf_path)

    print_group_3_4_5(table3, table4, table5, start, count, output_path, sheet_name)

    table6, table7 = get_dev_and_mes(table3)

    print_group_6_7(table6, table7, start, count, output_path, sheet_name)

    os.remove(excel_path)



pdf_path = input("Enter the file path: ")
output_path = "final.xlsx"
folder_path = input("Enter the folder path: ")
file_or_folder = 2

if file_or_folder == 1:
    perform_all_file_actions(pdf_path, output_path)
elif file_or_folder == 2:
    file_list = folder_list(folder_path)
    print(file_list)
    pdf_len = len(file_list)
    for i in range(pdf_len):
        pdf_path_folder = file_list[i]
        perform_all_file_actions(pdf_path_folder, output_path)
        print("Performig operation on file: " + pdf_path_folder)
else:
    print("Invalid input. Choose file or folder")
