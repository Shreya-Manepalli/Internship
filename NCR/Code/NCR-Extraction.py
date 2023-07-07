'''
######################################################## USING PyQt5 GUI ########################################################

import sys
from PyQt5.QtWidgets import QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton, QFileDialog
import tabula
import pandas as pd
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
import os

class MyWindow(QMainWindow):
    def __init__(self):
        super().__init__()

        self.input_file = None
        self.output_file = None

        self.setWindowTitle("Non-Conformance Report to Excel coverter")
        self.setGeometry(300, 100, 700, 200)
        self.setStyleSheet("background-color: #f0f0f0;")

        layout = QVBoxLayout()
        central_widget = QWidget(self)
        central_widget.setLayout(layout)
        self.setCentralWidget(central_widget)

        self.input_label = QLabel("Non Conformance Report(PDF Format):")
        layout.addWidget(self.input_label)

        self.input_button = QPushButton("Choose the Non Conformance Report (PDF Format)")
        self.input_button.setStyleSheet("background-color: #e0e0e0;")
        self.input_button.clicked.connect(self.choose_input_file)
        layout.addWidget(self.input_button)

        self.output_label = QLabel("Output Excel File:")
        layout.addWidget(self.output_label)

        self.output_button = QPushButton("Choose the Excel File")
        self.output_button.setStyleSheet("background-color: #e0e0e0;")
        self.output_button.clicked.connect(self.choose_output_file)
        layout.addWidget(self.output_button)

        self.process_button = QPushButton("Convert")
        self.process_button.setStyleSheet("background-color: #e0e0e0;")
        self.process_button.clicked.connect(self.process_files)
        layout.addWidget(self.process_button)

    def choose_input_file(self):
        file_dialog = QFileDialog()
        self.input_file, _ = file_dialog.getOpenFileName(self, "Choose Input PDF File")
        self.input_label.setText(f"Input PDF File: {self.input_file}")

    def choose_output_file(self):
        file_dialog = QFileDialog()
        self.output_file, _ = file_dialog.getSaveFileName(self, "Choose Output Excel File")
        self.output_label.setText(f"Output Excel File: {self.output_file}")

    def process_files(self):
        if self.input_file and self.output_file:
            pdf_path = self.input_file
            excel_path = "out1.xlsx"
            output_path = self.output_file

            def write_to_excel(string, row_no, col_no, file_path):
                try:
                    workbook = load_workbook(file_path)
                except FileNotFoundError:
                    workbook = Workbook()

                sheet = workbook.active
                sheet.cell(row=row_no, column=col_no).value = string
                workbook.save(file_path)

            dfs = tabula.read_pdf(pdf_path, pages='all')
            with pd.ExcelWriter(excel_path) as writer:
                for i, df in enumerate(dfs):
                    df.to_excel(writer, sheet_name=f"Sheet{i+1}", index=False)

            workbook = load_workbook(excel_path)
            sheet_names = workbook.sheetnames

            total_sheets = len(sheet_names)

            cells_table1 = [[2,4], [4,1], [6,3]]
            total_cells_table1 = len(cells_table1)

            cells_table2 = [[1,3], [1,4], [2,1], [2,1]]
            total_cells_table2 = len(cells_table2)

            table1 = []

            empty_rows = total_sheets
            empty_cols = 4

            table2 = [[None for _ in range(empty_cols)] for _ in range(empty_rows)]

            i = 0
            j = 0

            count = 0
            currentRow = 0
            select = 0

            for i in range(total_sheets):

                sheet = workbook[sheet_names[i]]

                if i == 0:
                    for j in range(total_cells_table1):
                        row_number = cells_table1[j][0]
                        column_number = cells_table1[j][1]
                        cell_value = sheet.cell(row=row_number, column=column_number).value
                        data = cell_value
                        table1.append(data)


                if i in range(2, total_sheets):

                    if count == 0:
                        count = count + 1

                    elif count == 1:

                        row_number = cells_table2[select][0]
                        column_number = cells_table2[select][1]
                        cell_value = sheet.cell(row=row_number, column=column_number).value
                        data = cell_value
                        temp1 = data[18:]
                        table2[currentRow][0] = temp1

                        select = select + 1

                        row_number = cells_table2[select][0]
                        column_number = cells_table2[select][1]
                        cell_value = sheet.cell(row=row_number, column=column_number).value
                        data = cell_value
                        temp2 = data[17:]
                        table2[currentRow][1] = temp2

                        select = select + 1
                        count = count + 1

                    elif count == 2:
                        row_number = cells_table2[select][0]
                        column_number = cells_table2[select][1]
                        cell_value = sheet.cell(row=row_number, column=column_number).value
                        data = cell_value
                        temp3 = data
                        table2[currentRow][2] = temp3

                        select = select + 1
                        count = count + 1

                    elif count == 3:
                        row_number = cells_table2[select][0]
                        column_number = cells_table2[select][1]
                        cell_value = sheet.cell(row=row_number, column=column_number).value
                        data = cell_value
                        temp4 = data
                        table2[currentRow][3] = temp4

                        select = 0
                        count = count + 1

                    elif count == 4:
                        count = count + 1
                        select = 0

                    elif count == 5:
                        count = 0
                        select = 0
                        currentRow = currentRow + 1

            write_to_excel("NCR-No.", 1, 1, output_path)
            write_to_excel("Material No.", 2, 1, output_path)
            write_to_excel("Design Organization Drawing and issue", 3, 1, output_path)

            z = 0

            for z in range(3):
                write_to_excel(table1[z], z+1, 2, output_path)

            write_to_excel("Item No.", 4, 1, output_path)
            write_to_excel("Defect Type", 4, 2, output_path)
            write_to_excel("Charact. No.", 4, 3, output_path)
            write_to_excel("Serial numbers affected", 4, 4, output_path)
            write_to_excel("Description of Nonconformance", 4, 5, output_path)

            y = 0
            x = 0
            item_no = 1

            for x in range(currentRow):

                excel_row = x + 5
                for y in range(5):

                    if y < 4:
                        excel_col = y + 2
                        write_to_excel(table2[x][y], excel_row, excel_col, output_path)

                    elif y == 4:
                        str_item_no = str(item_no)
                        write_to_excel(str_item_no, excel_row, 1, output_path)
                        item_no = item_no + 1

            os.remove(excel_path)
        else:
            print("Please choose both input and output files.")

if __name__ == '__main__':
    app = QApplication(sys.argv)
    window = MyWindow()
    window.show()
    sys.exit(app.exec_())


'''

######################################################## USING PySimpleGUI GUI ########################################################



import sys
import tabula
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
import os
import PySimpleGUI as sg

sg.theme('Purple')
SYMBOL_UP = '▲'
SYMBOL_DOWN = '▼'

def collapse(sectionLayout, visible, key):
    return sg.pin(sg.Column(sectionLayout, visible=visible, key=key))

# LAYOUT
layout = [
    [sg.Text('Select the Non Conformance Report(PDF Format):', size=(40,1), font=(None, 10, 'bold')),
     sg.InputText(key='pdf_path'), sg.FileBrowse()],
    [sg.Text('Choose the excel file:', size=(40, 1), font=(None, 10, 'bold')),
     sg.InputText(key='output_path'), sg.FileBrowse()],
    [sg.Button('Submit'), sg.Button('Exit')],
    [sg.Output(size=(120, 5))]
]

def process_files(input_file, output_file):
    if input_file and output_file:
        pdf_path = input_file
        excel_path = "out1.xlsx"
        output_path = output_file

        def write_to_excel(string, row_no, col_no, file_path):
            try:
                workbook = load_workbook(file_path)
            except FileNotFoundError:
                workbook = Workbook()

            sheet = workbook.active
            sheet.cell(row=row_no, column=col_no).value = string
            workbook.save(file_path)

        dfs = tabula.read_pdf(pdf_path, pages='all')
        with pd.ExcelWriter(excel_path) as writer:
            for i, df in enumerate(dfs):
                df.to_excel(writer, sheet_name=f"Sheet{i+1}", index=False)

        workbook = load_workbook(excel_path)
        sheet_names = workbook.sheetnames

        total_sheets = len(sheet_names)

        cells_table1 = [[2, 4], [4, 1], [6, 3]]
        total_cells_table1 = len(cells_table1)

        cells_table2 = [[1, 3], [1, 4], [2, 1], [2, 1]]
        total_cells_table2 = len(cells_table2)

        table1 = []

        empty_rows = total_sheets
        empty_cols = 4

        table2 = [[None for _ in range(empty_cols)] for _ in range(empty_rows)]

        i = 0
        j = 0

        count = 0
        currentRow = 0
        select = 0

        for i in range(total_sheets):
            sheet = workbook[sheet_names[i]]

            if i == 0:
                for j in range(total_cells_table1):
                    row_number = cells_table1[j][0]
                    column_number = cells_table1[j][1]
                    cell_value = sheet.cell(row=row_number, column=column_number).value
                    data = cell_value
                    table1.append(data)

            if i in range(2, total_sheets):
                if count == 0:
                    count = count + 1
                elif count == 1:
                    row_number = cells_table2[select][0]
                    column_number = cells_table2[select][1]
                    cell_value = sheet.cell(row=row_number, column=column_number).value
                    data = cell_value
                    temp1 = data[18:]
                    table2[currentRow][0] = temp1
                    select = select + 1

                    row_number = cells_table2[select][0]
                    column_number = cells_table2[select][1]
                    cell_value = sheet.cell(row=row_number, column=column_number).value
                    data = cell_value
                    temp2 = data[17:]
                    table2[currentRow][1] = temp2
                    select = select + 1
                    count = count + 1
                elif count == 2:
                    row_number = cells_table2[select][0]
                    column_number = cells_table2[select][1]
                    cell_value = sheet.cell(row=row_number, column=column_number).value
                    data = cell_value
                    temp3 = data
                    table2[currentRow][2] = temp3
                    select = select + 1
                    count = count + 1
                elif count == 3:
                    row_number = cells_table2[select][0]
                    column_number = cells_table2[select][1]
                    cell_value = sheet.cell(row=row_number, column=column_number).value
                    data = cell_value
                    temp4 = data
                    table2[currentRow][3] = temp4
                    select = 0
                    count = count + 1
                elif count == 4:
                    count = count + 1
                    select = 0
                elif count == 5:
                    count = 0
                    select = 0
                    currentRow = currentRow + 1

        write_to_excel("NCR-No.", 1, 1, output_path)
        write_to_excel("Material No.", 2, 1, output_path)
        write_to_excel("Design Organization Drawing and issue", 3, 1, output_path)

        z = 0

        for z in range(3):
            write_to_excel(table1[z], z+1, 2, output_path)

        write_to_excel("Item No.", 4, 1, output_path)
        write_to_excel("Defect Type", 4, 2, output_path)
        write_to_excel("Charact. No.", 4, 3, output_path)
        write_to_excel("Serial numbers affected", 4, 4, output_path)
        write_to_excel("Description of Nonconformance", 4, 5, output_path)

        y = 0
        x = 0
        item_no = 1

        for x in range(currentRow):
            excel_row = x + 5
            for y in range(5):
                if y < 4:
                    excel_col = y + 2
                    write_to_excel(table2[x][y], excel_row, excel_col, output_path)
                elif y == 4:
                    str_item_no = str(item_no)
                    write_to_excel(str_item_no, excel_row, 1, output_path)
                    item_no = item_no + 1

        os.remove(excel_path)

window = sg.Window('Non Conformance Report to Excel Converter', layout)

while True:
    event, values = window.read()

    if event == sg.WIN_CLOSED or event == 'Exit':
        break

    if event == 'Submit':
        print('Running............')
        input_file = values['pdf_path']
        output_file = values['output_path']
        process_files(input_file, output_file)
        print('Done!!!!')

    if event.startswith('-OPEN SEC1-'):
        opened1 = not opened1
        window['-OPEN SEC1-'].update(SYMBOL_UP if opened1 else SYMBOL_DOWN)
        window['-SEC1-'].update(visible=opened1)

window.close()




