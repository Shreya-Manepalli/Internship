import tabula
import pandas as pd
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
import os
import glob


def folder_list(folder_path):
    files_list = []

    pdf_files = glob.glob(f"{folder_path}/*.pdf")
    if pdf_files:
        for file_path in pdf_files:
            data = file_path
            files_list.append(data)
            #print(data)

    else:
        print("No PDF files found in the folder.")

    return files_list


def find_last_row(file_path, sheet_name, column_index):
    try:
        workbook = load_workbook(file_path)
    except FileNotFoundError:
        return 0
    
    # Check if the sheet exists
    if sheet_name not in workbook.sheetnames:
        return 0

    sheet = workbook[sheet_name]
    max_row = sheet.max_row
    
    for row in range(max_row, 0, -1):
        cell_value = sheet.cell(row=row, column=column_index).value
        if cell_value is not None:
            return row

    return 0  # If no filled rows are found



def write_to_excel(string, row_no, col_no, file_path):
    try:
        workbook = load_workbook(file_path)
    except FileNotFoundError:
        workbook = Workbook()

    sheet = workbook.active
    sheet.cell(row=row_no, column=col_no).value = string
    workbook.save(file_path)



def pdf_2_excel(pdf_path, excel_path, output_path):
    dfs = tabula.read_pdf(pdf_path, pages='all')
    with pd.ExcelWriter(excel_path) as writer:
        for i, df in enumerate(dfs):
            df.to_excel(writer, sheet_name=f"Sheet{i+1}", index=False)

    workbook = load_workbook(excel_path)
    sheet_names = workbook.sheetnames

    total_sheets = len(sheet_names)

    # NCR, Material, Design (in order)
    cells_table1 = [[2,4], [4,1], [6,3]]
    total_cells_table1 = len(cells_table1)

    # (DefectType, CharNo)4, (SerialNos)5, (Description)6
    cells_table2 = [[1,3], [1,4], [2,1], [2,1]]
    total_cells_table2 = len(cells_table2)

    table1 = []

    empty_rows = total_sheets  # Number of rows
    empty_cols = 4  # Number of columns

    # Create an empty 2D array
    table2 = [[None for _ in range(empty_cols)] for _ in range(empty_rows)]

    i = 0
    j = 0 

    count = 0
    currentRow = 0
    select = 0

    for i in range(total_sheets):
        
        sheet = workbook[sheet_names[i]]

        if i == 0:
            for j in range(total_cells_table1):
                row_number = cells_table1[j][0]
                column_number = cells_table1[j][1]
                cell_value = sheet.cell(row=row_number, column=column_number).value
                data = cell_value
                table1.append(data)
        
        if i in range(2, total_sheets):

            if count == 0:
                count = count + 1
            
            elif count == 1:

                row_number = cells_table2[select][0]
                column_number = cells_table2[select][1]
                cell_value = sheet.cell(row=row_number, column=column_number).value
                data = cell_value
                temp1 = data[18:]
                table2[currentRow][0] = temp1

                select = select + 1

                row_number = cells_table2[select][0]
                column_number = cells_table2[select][1]
                cell_value = sheet.cell(row=row_number, column=column_number).value
                data = cell_value
                temp2 = data[17:]
                table2[currentRow][1] = temp2

                select = select + 1
                count = count + 1

            elif count == 2:
                row_number = cells_table2[select][0]
                column_number = cells_table2[select][1]
                cell_value = sheet.cell(row=row_number, column=column_number).value
                data = cell_value
                temp3 = data
                table2[currentRow][2] = temp3

                select = select + 1
                count = count + 1

            elif count == 3:
                row_number = cells_table2[select][0]
                column_number = cells_table2[select][1]
                cell_value = sheet.cell(row=row_number, column=column_number).value
                data = cell_value
                temp4 = data
                table2[currentRow][3] = temp4

                select = 0
                count = count + 1

            elif count == 4:
                count = count + 1
                select = 0

            elif count == 5:
                count = 0
                select = 0
                currentRow = currentRow + 1


    output_path = 'result.xlsx'
    output_sheet_name = 'Sheet'
    column_index = 1

    start_point = 0

    last_filled_row = find_last_row(output_path, output_sheet_name, column_index)

    if last_filled_row == 0:
        start_point = last_filled_row 
    else:
        start_point = last_filled_row + 3

    # Setting up workbook
    write_to_excel("Material No.", 1 + start_point, 1, output_path)
    write_to_excel("NCR-No.", 2 + start_point, 1, output_path)
    write_to_excel("Design Organization Drawing and issue", 3 + start_point, 1, output_path)

    z = 0

    for z in range(3):
        write_to_excel(table1[z], z+1 + start_point, 2, output_path)

    write_to_excel("Item No.", 4 + start_point, 1, output_path)
    write_to_excel("Defect Type", 4 + start_point, 2, output_path)
    write_to_excel("Charact. No.", 4 + start_point, 3, output_path)
    write_to_excel("Serial numbers affected", 4 + start_point, 4, output_path)
    write_to_excel("Description of Nonconformance", 4 + start_point, 5, output_path)

    y = 0
    x = 0
    item_no = 1

    for x in range(currentRow):

        excel_row = x + 5

        for y in range(5):
            
            if y < 4:
                excel_col = y + 2
                write_to_excel(table2[x][y], excel_row + start_point, excel_col, "result.xlsx")

            
            elif y == 4:
                str_item_no = str(item_no)
                write_to_excel(str_item_no, excel_row + start_point, 1, "result.xlsx")
                item_no = item_no + 1
    try:
        os.remove(excel_path)
        print(f"File '{excel_path}' has been deleted.")
    except FileNotFoundError:
        print(f"File '{excel_path}' not found.")
    except PermissionError:
        print(f"Permission denied to delete file '{excel_path}'.")
  



folder_path = input("Enter the foler path: ")

pdf_path = "NCR-edited.pdf"

excel_path = "out2.xlsx"

output_path = "result.xlsx"

# Type 1 to enter a file and 2 to enter a folder: 

file_or_folder = 2

if file_or_folder == 1:
    pdf_2_excel(pdf_path, excel_path, output_path)
elif file_or_folder == 2:
    file_list = folder_list(folder_path)
    pdf_len = len(file_list)
    i = 0
    for i in range(pdf_len):
        pdf_path_folder = file_list[i]
        pdf_2_excel(pdf_path_folder, excel_path, output_path)
else:
    print("Invalid input")


