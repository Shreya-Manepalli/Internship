import tabula
import pandas as pd
from openpyxl import load_workbook

pdf_path = "sample2.pdf"
excel_path = "path/to/out2.xlsx"

dfs = tabula.read_pdf(pdf_path, pages='all')
with pd.ExcelWriter(excel_path) as writer:
    for i, df in enumerate(dfs):
        df.to_excel(writer, sheet_name=f"Sheet{i+1}", index=False)

workbook = load_workbook(excel_path)
sheet_names = workbook.sheetnames

total_sheets = len(sheet_names)

# NCR, Material, Design (in order)
cells_table1 = [[2,4], [4,1], [6,3]]
total_cells_table1 = len(cells_table1)

# (DefectType, CharNo)4, (SerialNos)5, (Description)6
cells_table2 = [[1,3], [1,4], [2,1], [2,1]]
total_cells_table2 = len(cells_table2)

table1 = []
table2 = [[]]

i = 0
j = 0 

count = 0
currentRow = 1
select = 0

for i in range(total_sheets):
    
    sheet = workbook[sheet_names[i]]

    if i == 0:
        for j in range(total_cells_table1):
            row_number = cells_table1[j][0]
            column_number = cells_table1[j][1]
            cell_value = sheet.cell(row=row_number, column=column_number).value
            data = cell_value
            table1.append(data)
        print(table1)

    # print("------" + str(sheet) + "------")
    
    if i in range(2, total_sheets):

        if count == 0:
            count = count + 1
        
        elif count == 1:

            print("====DefectType====")

            row_number = cells_table2[select][0]
            column_number = cells_table2[select][1]
            cell_value = sheet.cell(row=row_number, column=column_number).value
            data = cell_value
            print(data[12:]) # + "(" + str(row_number) + ", " + str(column_number) + ")" + " Select = " + str(select))

            select = select + 1

            print("====CharNo====")

            row_number = cells_table2[select][0]
            column_number = cells_table2[select][1]
            cell_value = sheet.cell(row=row_number, column=column_number).value
            data = cell_value
            print(data[11:]) # + "(" + str(row_number) + ", " + str(column_number) + ")" + " Select = " + str(select))

            select = select + 1
            count = count + 1

        elif count == 2:

            print("====SerialNos====")

            row_number = cells_table2[select][0]
            column_number = cells_table2[select][1]
            cell_value = sheet.cell(row=row_number, column=column_number).value
            data = cell_value
            print(data) # + "(" + str(row_number) + ", " + str(column_number) + ")" + " Select = " + str(select))

            select = select + 1
            count = count + 1

        elif count == 3:

            print("====Description====")
            
            row_number = cells_table2[select][0]
            column_number = cells_table2[select][1]
            cell_value = sheet.cell(row=row_number, column=column_number).value
            data = cell_value
            print(data) # + "(" + str(row_number) + ", " + str(column_number) + ")" + " Select = " + str(select))

            select = 0
            count = count + 1

        elif count == 4:
            count = count + 1
            select = 0

        elif count == 5:
            print("===========================================================")
            count = 0
            select = 0

    








