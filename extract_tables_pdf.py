import locale
import PySimpleGUI as sg
import tabula
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles.borders import Border, Side
import pandas as pd

sg.theme('Purple')
SYMBOL_UP = '▲'
SYMBOL_DOWN = '▼'

def collapse(sectionLayout, visible, key):
    return sg.pin(sg.Column(sectionLayout, visible=visible, key=key))

# ---------------MAIN LAYOUT---------------
layout = [
    [sg.Text('Thousand & Decimal separator in PDF document:', font=(None, 9, 'bold'))],
    [sg.Radio('1.234.567,999', 'RADIO1', default=True, size=(19, 1), key='locale'),
     sg.Radio('1,234,567.999', 'RADIO1', default=False, size=(19, 1))],

    [sg.Text('Source for PDF document:', size=(25, 1), font=(None, 9, 'bold')),
     sg.InputText(key='pdfDocument'), sg.FileBrowse()],

    [sg.Text('Source of Excel file:', size=(25, 1), font=(None, 9, 'bold'))],
    [sg.Radio('Existed Excel file:', 'RADIO0', size=(19, 1), default=True), sg.InputText(key='excelTemplate'),
     sg.FileBrowse()],
    [sg.Radio('Create a blank Excel file', 'RADIO0', default=False, key='newWb')],

    [sg.Text('Choose any of the following options:', font=(None, 9, 'bold'))],
    [sg.Checkbox('Apply cell border (All Borders)', default=False, key='border')],
    [sg.Checkbox('Copy text format (Bold, Italic & Underline)', default=False, key='formatText')],
    [sg.Checkbox('Convert cell "-" to 0', default=False, key='hyphenToZero')],

    [sg.Text('Choose folder to save:', size=(25, 1), font=(None, 9, 'bold')),
     sg.InputText(key='savedFolder'), sg.FolderBrowse()],
    [sg.Text('Enter file name (Ex: file.xlsx):', size=(25, 1), font=(None, 9, 'bold')),
     sg.InputText(default_text='file.xlsx', key='savedName')],

    [sg.Button('Submit'), sg.Button('Exit')],
    [sg.Output(size=(80, 3))]
]

# Modify the pdfToExcel function
def pdfToExcel(dict):
    # Load existed Excel file or create a new blank Excel file
    if not dict['newWb']:
        wb = openpyxl.load_workbook(dict['excelTemplate'])
    else:
        wb = Workbook()

    # Load the PDF file using Tabula
    tables = tabula.read_pdf(dict['pdfDocument'], pages='all', multiple_tables=True)

    if dict['locale']:
        locale.setlocale(locale.LC_NUMERIC, "en_DK.UTF-8")  # 1.234.567,999
    else:
        locale.setlocale(locale.LC_NUMERIC, "en_US.UTF-8")  # 1,234,567.999

    # Set border style for Excel cells
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    # Loop through each table extracted from the PDF
    for table_num, table in enumerate(tables):
        # Create a new sheet for each table
        sheet = wb.create_sheet(title=f'Table {table_num + 1}')
        excel_writer = pd.ExcelWriter('tables.xlsx', engine='xlsxwriter')

    for i, table in enumerate(tables):
    # Write each DataFrame to a separate sheet
        sheet_name = f"Table {i+1}"
        table.to_excel(excel_writer, sheet_name=sheet_name, index=False)

    excel_writer.save()

        # Loop through each row in the table
    for i, row in enumerate(tables):
            # Loop through each cell in the row
            for j, cell in enumerate(row):
                # Excel cell starts at (1, 1)
                excelCell = sheet.cell(row=i + 1, column=j + 1)

                # Convert number from PDF to number in Excel
                try:
                    excelCell.value = locale.atof(cell.replace('(', '-').replace(')', ''))
                except ValueError:
                    if dict['hyphenToZero'] and cell == "-":
                        excelCell.value = 0
                    else:
                        excelCell.value = cell

                # Apply cell border
                if dict['border']:
                    excelCell.border = thin_border

                # Copy text format
                if dict['formatText']:
                    excelCell.font = Font(bold=False,
                                          italic=False,
                                          underline=False)

    # Save the Excel file
    if not dict['savedName'].endswith('.xlsx'):
        dict['savedName'] += '.xlsx'
    wb.save((dict['savedFolder'] + '/' + dict['savedName']))

# load GUI
window = sg.Window('Copy PDF tables to Excel', layout)

while True:  # Event Loop
    event, values = window.read()
    if event == sg.WIN_CLOSED or event == 'Exit':
        break
    if event == 'Submit':
        # Print message to Output box
        print('Running............')

        # convert the 'window dict' to 'normal dict' because 'window dict' values have different class(es)
        keys = ['excelTemplate', 'newWb', 'pdfDocument',
                'locale',
                'border', 'hyphenToZero', 'formatText',
                'savedFolder', 'savedName'
                ]
        convertDict = {x: window[x].get() for x in keys}

        # run 'pdfToExcel'
        if convertDict['pdfDocument'] and convertDict['savedFolder']:
            pdfToExcel(convertDict)

        # Print message to Output box
        print('Done!!!!')

    if event.startswith('-OPEN SEC1-'):
        opened1 = not opened1
        window['-OPEN SEC1-'].update(SYMBOL_UP if opened1 else SYMBOL_DOWN)
        window['-SEC1-'].update(visible=opened1)

window.close()

