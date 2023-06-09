import tabula
import pandas as pd
from openpyxl import load_workbook

pdf_path = "sample2.pdf"
excel_path = "out2.xlsx"

dfs = tabula.read_pdf(pdf_path, pages='all')
with pd.ExcelWriter(excel_path) as writer:
    for i, df in enumerate(dfs):
        df.to_excel(writer, sheet_name=f"Sheet{i+1}", index=False)

workbook = load_workbook(excel_path)
sheet_names = workbook.sheetnames

combined_data = []

for sheet_name in sheet_names:
    sheet = workbook[sheet_name]

    rows_needed = [[2, 1], [2, 2], [4, 2], [3, 4], [5, 4], [6, 2], [6, 4]]
    data = {}

    for row_number, column_number in rows_needed:
        cell_value = sheet.cell(row=row_number, column=column_number).value

        if row_number == 2 and column_number == 1:
            data['Author'] = cell_value[14:]
        elif row_number == 2 and column_number == 2:
            data['Document No'] = cell_value[23:]
        elif row_number == 4 and column_number == 2:
            data['Company'] = cell_value[15:]
        elif row_number == 3 and column_number == 4:
            data['Date'] = cell_value
        elif row_number == 5 and column_number == 4:
            data['Smile Approved By'] = cell_value
        elif row_number == 6 and column_number == 2:
            data['Review By'] = cell_value
        elif row_number == 6 and column_number == 4:
            data['Review Date'] = cell_value

    combined_data.append(data)

df = pd.DataFrame(combined_data)
print(df)
df.to_csv('file2.csv', index=False)
