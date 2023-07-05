'''
import tabula
import pandas as pd
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
import os
import glob

def write_to_excel(string, row_no, col_no, file_path):
    try:
        workbook = load_workbook(file_path)
    except FileNotFoundError:
        workbook = Workbook()

    sheet = workbook.active
    sheet.cell(row=row_no, column=col_no).value = string
    workbook.save(file_path)

pdf_folder = "NCR"
excel_folder = "Output"

pdf_files = glob.glob(os.path.join(pdf_folder, "*.pdf"))

if not os.path.exists(excel_folder):
    os.makedirs(excel_folder)

for pdf_file in pdf_files:
    pdf_filename = os.path.basename(pdf_file)
    excel_filename = os.path.splitext(pdf_filename)[0] + ".xlsx"

    pdf_path = os.path.join(pdf_folder, pdf_filename)
    excel_path = os.path.join(excel_folder, excel_filename)

    dfs = tabula.read_pdf(pdf_path, pages='all')
    with pd.ExcelWriter(excel_path) as writer:
        for i, df in enumerate(dfs):
            df.to_excel(writer, sheet_name=f"Sheet{i+1}", index=False)

    workbook = load_workbook(excel_path)
    sheet_names = workbook.sheetnames

    total_sheets = len(sheet_names)

    cells_table1 = [[2,4], [4,1], [6,3]]
    total_cells_table1 = len(cells_table1)

    cells_table2 = [[1,3], [1,4], [2,1], [2,1]]
    total_cells_table2 = len(cells_table2)

    table1 = []
    empty_rows = total_sheets
    empty_cols = 4

    table2 = [[None for _ in range(empty_cols)] for _ in range(empty_rows)]

    i = 0
    j = 0
    count = 0
    currentRow = 0
    select = 0

    for i in range(total_sheets):
        sheet = workbook[sheet_names[i]]

        if i == 0:
            for j in range(total_cells_table1):
                row_number = cells_table1[j][0]
                column_number = cells_table1[j][1]
                cell_value = sheet.cell(row=row_number, column=column_number).value
                data = cell_value
                table1.append(data)

        if i in range(2, total_sheets):
            if count == 0:
                count = count + 1
            elif count == 1:
                row_number = cells_table2[select][0]
                column_number = cells_table2[select][1]
                cell_value = sheet.cell(row=row_number, column=column_number).value
                data = cell_value
                temp1 = data[12:]
                table2[currentRow][0] = temp1

                select = select + 1

                row_number = cells_table2[select][0]
                column_number = cells_table2[select][1]
                cell_value = sheet.cell(row=row_number, column=column_number).value
                data = cell_value
                temp2 = data[11:]
                table2[currentRow][1] = temp2

                select = select + 1
                count = count + 1
            elif count == 2:
                row_number = cells_table2[select][0]
                column_number = cells_table2[select][1]
                cell_value = sheet.cell(row=row_number, column=column_number).value
                data = cell_value
                temp3 = data
                table2[currentRow][2] = temp3

                select = select + 1
                count = count + 1
            elif count == 3:
                row_number = cells_table2[select][0]
                column_number = cells_table2[select][1]
                cell_value = sheet.cell(row=row_number, column=column_number).value
                data = cell_value
                temp4 = data
                table2[currentRow][3] = temp4

                select = 0
                count = count + 1
            elif count == 4:
                count = count + 1
                select = 0
            elif count == 5:
                count = 0
                select = 0
                currentRow = currentRow + 1

    write_to_excel("Material No.", 1, 1, excel_path)
    write_to_excel("NCR-No.", 2, 1, excel_path)
    write_to_excel("Design Organization Drawing and issue", 3, 1, excel_path)

    z = 0

    for z in range(3):
        write_to_excel(table1[z], z+1, 2, excel_path)

    write_to_excel("Item No.", 4, 1, excel_path)
    write_to_excel("Defect Type", 4, 2, excel_path)
    write_to_excel("Charact. No.", 4, 3, excel_path)
    write_to_excel("Serial numbers affected", 4, 4, excel_path)
    write_to_excel("Description of Nonconformance", 4, 5, excel_path)

    y = 0
    x = 0
    item_no = 1

    for x in range(currentRow):
        excel_row = x + 5
        for y in range(5):
            if y < 4:
                excel_col = y + 2
                write_to_excel(table2[x][y], excel_row, excel_col, excel_path)
            elif y == 4:
                str_item_no = str(item_no)
                write_to_excel(str_item_no, excel_row, 1, excel_path)
                item_no = item_no + 1
'''

import tabula
import pandas as pd
from openpyxl import load_workbook

import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
import os

pdf_path = input("Enter the path of the pdf file: ")
excel_path = "out2.xlsx"
output_path = input("Enter the name of the excel file(example: result.xlsx): ")


def write_to_excel(string, row_no, col_no, file_path):
    try:
        workbook = load_workbook(file_path)
    except FileNotFoundError:
        workbook = Workbook()

    sheet = workbook.active
    sheet.cell(row=row_no, column=col_no).value = string
    workbook.save(file_path)

dfs = tabula.read_pdf(pdf_path, pages='all')
with pd.ExcelWriter(excel_path) as writer:
    for i, df in enumerate(dfs):
        df.to_excel(writer, sheet_name=f"Sheet{i+1}", index=False)

workbook = load_workbook(excel_path)
sheet_names = workbook.sheetnames

total_sheets = len(sheet_names)


cells_table1 = [[2,4], [4,1], [6,3]]
total_cells_table1 = len(cells_table1)


cells_table2 = [[1,3], [1,4], [2,1], [2,1]]
total_cells_table2 = len(cells_table2)

table1 = []

empty_rows = total_sheets  
empty_cols = 4  

table2 = [[None for _ in range(empty_cols)] for _ in range(empty_rows)]

i = 0
j = 0 

count = 0
currentRow = 0
select = 0

for i in range(total_sheets):
    
    sheet = workbook[sheet_names[i]]

    if i == 0:
        for j in range(total_cells_table1):
            row_number = cells_table1[j][0]
            column_number = cells_table1[j][1]
            cell_value = sheet.cell(row=row_number, column=column_number).value
            data = cell_value
            table1.append(data)
        
    
    if i in range(2, total_sheets):

        if count == 0:
            count = count + 1
        
        elif count == 1:

            row_number = cells_table2[select][0]
            column_number = cells_table2[select][1]
            cell_value = sheet.cell(row=row_number, column=column_number).value
            data = cell_value
            temp1 = data[12:]
            table2[currentRow][0] = temp1

            select = select + 1

            row_number = cells_table2[select][0]
            column_number = cells_table2[select][1]
            cell_value = sheet.cell(row=row_number, column=column_number).value
            data = cell_value
            temp2 = data[11:]
            table2[currentRow][1] = temp2

            select = select + 1
            count = count + 1

        elif count == 2:
            row_number = cells_table2[select][0]
            column_number = cells_table2[select][1]
            cell_value = sheet.cell(row=row_number, column=column_number).value
            data = cell_value
            temp3 = data
            table2[currentRow][2] = temp3

            select = select + 1
            count = count + 1

        elif count == 3:
            row_number = cells_table2[select][0]
            column_number = cells_table2[select][1]
            cell_value = sheet.cell(row=row_number, column=column_number).value
            data = cell_value
            temp4 = data
            table2[currentRow][3] = temp4

            select = 0
            count = count + 1

        elif count == 4:
            count = count + 1
            select = 0

        elif count == 5:
            count = 0
            select = 0
            currentRow = currentRow + 1

write_to_excel("Material No.", 1, 1, output_path)
write_to_excel("NCR-No.", 2, 1, output_path)
write_to_excel("Design Organization Drawing and issue", 3, 1, output_path)

z = 0

for z in range(3):
    write_to_excel(table1[z], z+1, 2, output_path)

write_to_excel("Item No.", 4, 1, output_path)
write_to_excel("Defect Type", 4, 2, output_path)
write_to_excel("Charact. No.", 4, 3, output_path)
write_to_excel("Serial numbers affected", 4, 4, output_path)
write_to_excel("Description of Nonconformance", 4, 5, output_path)

y = 0
x = 0
item_no = 1

for x in range(currentRow):

    excel_row = x + 5
    for y in range(5):
        
        if y < 4:
            excel_col = y + 2
            write_to_excel(table2[x][y], excel_row, excel_col, output_path)

        
        elif y == 4:
            str_item_no = str(item_no)
            write_to_excel(str_item_no, excel_row, 1, output_path)
            item_no = item_no + 1

os.remove(excel_path)