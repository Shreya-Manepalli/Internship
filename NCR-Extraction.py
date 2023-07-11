######################################################## USING PySimpleGUI GUI ########################################################

import PyPDF2
import re
import sys
import tabula
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
import os
import PySimpleGUI as sg

sg.theme('Purple')
SYMBOL_UP = '▲'
SYMBOL_DOWN = '▼'

# Function to collapse a section in the GUI layout
def collapse(sectionLayout, visible, key):
    return sg.pin(sg.Column(sectionLayout, visible=visible, key=key))

# LAYOUT
layout = [
    [sg.Text('Select the Non Conformance Report(PDF Format):', size=(40,1), font=(None, 10, 'bold')),
     sg.InputText(key='pdf_path'), sg.FileBrowse()],
    [sg.Text('Choose the excel file:', size=(40, 1), font=(None, 10, 'bold')),
     sg.InputText(key='output_path'), sg.FileBrowse()],
    [sg.Button('Submit'), sg.Button('Exit')],
    [sg.Output(size=(120, 5))]
]

# Function to process files
def process_files(input_file, output_file, item_matches):
    if input_file and output_file:
        pdf_path = input_file
        excel_path = "out1.xlsx"
        output_path = output_file

        # Function to write data to an Excel file
        def write_to_excel(string, row_no, col_no, file_path):
            try:
                workbook = load_workbook(file_path)
            except FileNotFoundError:
                workbook = Workbook()

            sheet = workbook.active
            sheet.cell(row=row_no, column=col_no).value = string
            workbook.save(file_path)

        # Read PDF into DataFrame using tabula
        dfs = tabula.read_pdf(pdf_path, pages='all')
        with pd.ExcelWriter(excel_path) as writer:
            for i, df in enumerate(dfs):
                df.to_excel(writer, sheet_name=f"Sheet{i+1}", index=False)

        workbook = load_workbook(excel_path)
        sheet_names = workbook.sheetnames

        total_sheets = len(sheet_names)

        cells_table1 = [[2, 4], [4, 1], [6, 3]]
        total_cells_table1 = len(cells_table1)

        cells_table2 = [[1, 3], [1, 4], [2, 1], [2, 1]]
        total_cells_table2 = len(cells_table2)

        table1 = []

        empty_rows = total_sheets
        empty_cols = 4

        table2 = [[None for _ in range(empty_cols)] for _ in range(empty_rows)]

        i = 0
        j = 0

        count = 0
        currentRow = 0
        select = 0

        # Iterate through each sheet in the Excel file
        for i in range(total_sheets):
            sheet = workbook[sheet_names[i]]

            if i == 0:
                # Extract data from table 1
                for j in range(total_cells_table1):
                    row_number = cells_table1[j][0]
                    column_number = cells_table1[j][1]
                    cell_value = sheet.cell(row=row_number, column=column_number).value
                    data = cell_value
                    table1.append(data)

            if i in range(2, total_sheets):
                if count == 0:
                    count = count + 1
                elif count == 1:
                    # Extract data from table 2
                    row_number = cells_table2[select][0]
                    column_number = cells_table2[select][1]
                    cell_value = sheet.cell(row=row_number, column=column_number).value
                    data = cell_value
                    temp1 = data[18:]
                    table2[currentRow][0] = temp1
                    select = select + 1

                    row_number = cells_table2[select][0]
                    column_number = cells_table2[select][1]
                    cell_value = sheet.cell(row=row_number, column=column_number).value
                    data = cell_value
                    temp2 = data[17:]
                    table2[currentRow][1] = temp2
                    select = select + 1
                    count = count + 1
                elif count == 2:
                    row_number = cells_table2[select][0]
                    column_number = cells_table2[select][1]
                    cell_value = sheet.cell(row=row_number, column=column_number).value
                    data = cell_value
                    temp3 = data
                    table2[currentRow][2] = temp3
                    select = select + 1
                    count = count + 1
                elif count == 3:
                    row_number = cells_table2[select][0]
                    column_number = cells_table2[select][1]
                    cell_value = sheet.cell(row=row_number, column=column_number).value
                    data = cell_value
                    temp4 = data
                    table2[currentRow][3] = temp4
                    select = 0
                    count = count + 1
                elif count == 4:
                    count = count + 1
                    select = 0
                elif count == 5:
                    count = 0
                    select = 0
                    currentRow = currentRow + 1

        # Write table 1 data to Excel
        write_to_excel("NCR-No.", 1, 1, output_path)
        write_to_excel("Material No.", 2, 1, output_path)
        write_to_excel("Design Organization Drawing and issue", 3, 1, output_path)

        z = 0

        for z in range(3):
            write_to_excel(table1[z], z+1, 2, output_path)

        # Write table 2 headers to Excel
        write_to_excel("Item No.", 4, 1, output_path)
        write_to_excel("Defect Type", 4, 2, output_path)
        write_to_excel("Charact. No.", 4, 3, output_path)
        write_to_excel("Serial numbers affected", 4, 4, output_path)
        write_to_excel("Description of Nonconformance", 4, 5, output_path)

        y = 0
        x = 0

        # Write table 2 data to Excel with item_matches
        for x in range(currentRow):
            excel_row = x + 5
            for y in range(5):
                if y < 4:
                    excel_col = y + 2
                    write_to_excel(table2[x][y], excel_row, excel_col, output_path)
                elif y == 4:
                    if x < len(item_matches):
                        str_item_no = item_matches[x]
                    else:
                        str_item_no = ""
                    write_to_excel(str_item_no, excel_row, 1, output_path)

        os.remove(excel_path)

window = sg.Window('Non Conformance Report to Excel Converter', layout)

while True:
    event, values = window.read()

    if event == sg.WIN_CLOSED or event == 'Exit':
        break

    if event == 'Submit':
        print('Running............')
        input_file = values['pdf_path']
        output_file = values['output_path']

        # Extract "Item" occurrences from PDF
        rd = PyPDF2.PdfReader(input_file)
        item_matches = []
        for page in rd.pages:
            text = page.extract_text()
            matches = re.findall(r"Item No. \d", text)
            item_matches.extend(matches)

        # Call the modified process_files function with the extracted item_matches
        process_files(input_file, output_file, item_matches)
        print('Done!!!!')

    if event.startswith('-OPEN SEC1-'):
        opened1 = not opened1
        window['-OPEN SEC1-'].update(SYMBOL_UP if opened1 else SYMBOL_DOWN)
        window['-SEC1-'].update(visible=opened1)

window.close()
