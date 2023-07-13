import tabula
import pandas as pd

from openpyxl import Workbook
from openpyxl import load_workbook
import os
import glob
import PyPDF2
import re

# def find_numbers(string):
#     pattern = r'is\s+([\w\s.-]+)\s+or\s+([\w\s.-]+)'
#     match = re.search(pattern, string)
#     if match:
#         num1 = match.group(1).strip()
#         num2_match = re.search(r'\d+(?:\.\d+)?', match.group(2))
#         num2 = num2_match.group() if num2_match else None
#         return num1, num2
#     else:
#         return None, None
    
# def find_numbers(sentences):
#     pattern = r"is\s+(.*?)\s+or\s+(\d+\.?\d*)"
    
#     measured_values = []
#     deviation_values = []

#     for sentence in sentences:

#         if sentence.startswith("Casting Blend"):
#             measured_value = ""
#             deviation_value = ""
#         else:
#             match = re.search(pattern, sentence)
#             if match:
#                 measured_value = match.group(1)
#                 deviation_value = match.group(2)
#             else:
#                 measured_value = ""
#                 deviation_value = ""
#         measured_values.append(measured_value)
#         deviation_values.append(deviation_value)
#     return [measured_values, deviation_values]

def find_numbers(string):
    pattern = r'is\s+([\d.\s+-]+)[^\d.]*(\d+(?:\.\d+)?)'
    match = re.search(pattern, string)
    if match:
        num1 = match.group(1).strip()
        num2 = match.group(2).strip()
        return num1, num2
    else:
        return None, None
    
def get_item_no(pdf_path):
    rd = PyPDF2.PdfReader(pdf_path)
    item_matches = []
    for page in rd.pages:
        text = page.extract_text()
        matches = re.findall(r"Item No. \d", text)
        item_matches.extend(matches)
    return item_matches

def valid_sheet(string):
    characters_to_remove = ['*', ',', '.', '?', '/', '\\', ':', '[', ']']
    for char in characters_to_remove:
        string = string.replace(char, '')
    return string

def folder_list(folder_path):
    files_list = []

    pdf_files = glob.glob(f"{folder_path}/*.pdf")
    if pdf_files:
        for file_path in pdf_files:
            data = file_path
            files_list.append(data)
            # print(data)

    else:
        print("No PDF files found in the folder.")

    return files_list


def find_last_row(file_path, sheet_name, column_index):
    try:
        workbook = load_workbook(file_path)
    except FileNotFoundError:
        return 0
    
    # Check if the sheet exists
    if sheet_name not in workbook.sheetnames:
        return 0

    sheet = workbook[sheet_name]
    max_row = sheet.max_row
    
    for row in range(max_row, 0, -1):
        cell_value = sheet.cell(row=row, column=column_index).value
        if cell_value is not None:
            return row

    return 0  # If no filled rows are found



def write_to_excel(string, row_no, col_no, file_path, sheet_name):
    try:
        workbook = load_workbook(file_path)
    except FileNotFoundError:
        workbook = Workbook()

    if sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
    else:
        sheet = workbook.create_sheet(title=sheet_name)

    sheet.cell(row=row_no, column=col_no).value = string
    workbook.save(file_path)


def read_pdf(pdf_path, excel_path):
    dfs = tabula.read_pdf(pdf_path, pages='all')
    with pd.ExcelWriter(excel_path) as writer:
        for i, df in enumerate(dfs):
            df.to_excel(writer, sheet_name=f"Sheet{i+1}", index=False)

    workbook = load_workbook(excel_path)
    sheet_names = workbook.sheetnames

    total_sheets = len(sheet_names)

    # NCR, Material, Design (in order)
    cells_table1 = [[2,4], [4,1], [6,3]]
    total_cells_table1 = len(cells_table1)

    # (DefectType, CharNo)4, (SerialNos)5, (Description)6
    cells_table2 = [[1,3], [1,4], [2,1], [2,1]]
    total_cells_table2 = len(cells_table2)

    table1 = []

    empty_rows = total_sheets  # Number of rows
    empty_cols = 4  # Number of columns

    # Create an empty 2D array
    table2 = [[None for _ in range(empty_cols)] for _ in range(empty_rows)]

    i = 0
    j = 0 

    count = 0
    currentRow = 0
    select = 0

    for i in range(total_sheets):
        
        sheet = workbook[sheet_names[i]]

        if i == 0:
            for j in range(total_cells_table1):
                row_number = cells_table1[j][0]
                column_number = cells_table1[j][1]
                cell_value = sheet.cell(row=row_number, column=column_number).value
                data = cell_value
                table1.append(data)
            # print(table1)

        # print("------" + str(sheet) + "------")
        
        if i in range(2, total_sheets):

            if count == 0:
                count = count + 1
            
            elif count == 1:

                # print("====DefectType====")

                row_number = cells_table2[select][0]
                column_number = cells_table2[select][1]
                cell_value = sheet.cell(row=row_number, column=column_number).value
                data = cell_value
                # print(data[12:]) # + "(" + str(row_number) + ", " + str(column_number) + ")" + " Select = " + str(select))

                temp1 = data[12:]
                table2[currentRow][0] = temp1

                select = select + 1

                # print("====CharNo====")

                row_number = cells_table2[select][0]
                column_number = cells_table2[select][1]
                cell_value = sheet.cell(row=row_number, column=column_number).value
                data = cell_value
                # print(data[11:]) # + "(" + str(row_number) + ", " + str(column_number) + ")" + " Select = " + str(select))

                temp2 = data[11:]
                table2[currentRow][1] = temp2

                select = select + 1
                count = count + 1

            elif count == 2:

                # print("====SerialNos====")

                row_number = cells_table2[select][0]
                column_number = cells_table2[select][1]
                cell_value = sheet.cell(row=row_number, column=column_number).value
                data = cell_value
                # print(data) # + "(" + str(row_number) + ", " + str(column_number) + ")" + " Select = " + str(select))

                temp3 = data
                table2[currentRow][2] = temp3

                select = select + 1
                count = count + 1

            elif count == 3:

                # print("====Description====")
                
                row_number = cells_table2[select][0]
                column_number = cells_table2[select][1]
                cell_value = sheet.cell(row=row_number, column=column_number).value
                data = cell_value
                # print(data) # + "(" + str(row_number) + ", " + str(column_number) + ")" + " Select = " + str(select))

                temp4 = data
                table2[currentRow][3] = temp4

                select = 0
                count = count + 1

            elif count == 4:
                count = count + 1
                select = 0

            elif count == 5:
                # print("===========================================================")
                count = 0
                select = 0
                currentRow = currentRow + 1
    
    return [table1, table2, currentRow]

def table_2_excel(pdf_path, excel_path, output_path):

    [table1, table2, total_rows] = read_pdf(pdf_path, excel_path)

    output_path = output_path
    output_sheet_name = valid_sheet(str(table1[2]))
    column_index = 1

    start_point = 0

    last_filled_row = find_last_row(output_path, output_sheet_name, column_index)

    if last_filled_row == 0:
        start_point = last_filled_row 
    else:
        start_point = last_filled_row + 3

    # print("start point is: " + str(start_point))

    # Setting up workbook
    write_to_excel("Material No.", 1 + start_point, 2, output_path, output_sheet_name)
    write_to_excel("NCR-No.", 1 + start_point, 4, output_path, output_sheet_name)
    # write_to_excel("Design Organization Drawing and issue", 1 + start_point, 6, output_path, output_sheet_name)

    z = 0
    zz = 0

    for z in range(3):
        for zz in range(total_rows):
            if z == 0:
                excel_col = 4
                write_to_excel(table1[z], zz+2 + start_point, excel_col, output_path, output_sheet_name)
            if z == 1:
                excel_col = 2
                write_to_excel(table1[z], zz+2 + start_point, excel_col, output_path, output_sheet_name)
            if z == 2:
                continue
                # excel_col = 6    
                # write_to_excel(table1[z], zz+2 + start_point, excel_col, output_path, output_sheet_name)


    write_to_excel("Item No.", 1 + start_point, 1, output_path, output_sheet_name)
    write_to_excel("Defect Type", 1 + start_point, 6, output_path, output_sheet_name)
    write_to_excel("Charact. No.", 1 + start_point, 5, output_path, output_sheet_name)
    write_to_excel("Serial numbers affected", 1 + start_point, 3, output_path, output_sheet_name)
    write_to_excel("Description of Nonconformance", 1 + start_point, 7, output_path, output_sheet_name)
    write_to_excel("Measured Value", 1 + start_point, 8, output_path, output_sheet_name)
    write_to_excel("Deviation", 1 + start_point, 9, output_path, output_sheet_name)

    y = 0
    x = 0
    item_no = 1

    for x in range(total_rows):
        
        item_no_list = get_item_no(pdf_path)
        excel_row = x + 2
        
        # print("===========================================================")

        for y in range(5):
            
            if y < 4:

                if y == 0:
                    excel_col = 6
                    write_to_excel(table2[x][y], excel_row + start_point, excel_col, output_path, output_sheet_name)
                if y == 1:
                    excel_col = 5
                    write_to_excel(table2[x][y], excel_row + start_point, excel_col, output_path, output_sheet_name)
                if y == 2:
                    excel_col = 3
                    write_to_excel(table2[x][y], excel_row + start_point, excel_col, output_path, output_sheet_name)
                if y == 3:
                    # excel_col = 7
                    # write_to_excel(table2[x][y], excel_row + start_point, excel_col, output_path, output_sheet_name)
                    # [measured_values, deviation_values] = find_numbers([str(table2[x][y])])
                    # excel_col = 8
                    # write_to_excel(str(measured_values[0]), excel_row + start_point, excel_col, output_path, output_sheet_name)
                    # excel_col = 9
                    # write_to_excel(str(deviation_values[0]), excel_row + start_point, excel_col, output_path, output_sheet_name)
                    excel_col = 7
                    write_to_excel(table2[x][y], excel_row + start_point, excel_col, output_path, output_sheet_name)
                    measured_values, deviation_values = find_numbers(str(table2[x][y]))
                    excel_col = 8
                    write_to_excel(str(measured_values), excel_row + start_point, excel_col, output_path, output_sheet_name)
                    excel_col = 9
                    write_to_excel(str(deviation_values), excel_row + start_point, excel_col, output_path, output_sheet_name)
                    
                # excel_col = y + 5
                # print("=========================")
                # print(table2[x][y])

            
            elif y == 4:
                # print(item_no_list)
                excel_col = 1
                str_item_no = item_no_list[x]
                write_to_excel(str_item_no, excel_row + start_point, excel_col, output_path, output_sheet_name)
                item_no = item_no + 1
                
                

    try:
        os.remove(excel_path)
        # print(f"File '{excel_path}' has been deleted.")
    except FileNotFoundError:
        print(f"File '{excel_path}' not found.")
    except PermissionError:
        print(f"Permission denied to delete file '{excel_path}'.")


def pdf_file_2_excel(pdf_path, excel_path, output_path):
    temp_folder_path = "./path_temp"
    if not os.path.exists(temp_folder_path):
        os.makedirs(temp_folder_path)
        print(f"Folder '{temp_folder_path}' created.")


        table_2_excel(pdf_path, excel_path, output_path)


        os.rmdir(temp_folder_path)
        print(f"Folder '{temp_folder_path}' deleted.")
    else:
        print(f"Folder '{folder_path}' already exists.")
        table_2_excel(pdf_path, excel_path, output_path)



def pdf_folder_2_excel(folder_path, output_path, file_list):


    temp_folder_path = "./path_temp"

    pdf_len = len(file_list)
    i = 0

    if not os.path.exists(temp_folder_path):
        os.makedirs(temp_folder_path)
        # print(f"Folder '{temp_folder_path}' created.")

        pdf_len = len(file_list)
        i = 0
        for i in range(pdf_len):
            pdf_path_folder = file_list[i]
            table_2_excel(pdf_path_folder, excel_path, output_path)


        os.rmdir(temp_folder_path)
        # print(f"Folder '{temp_folder_path}' deleted.")
    else:
        # print(f"Folder '{temp_folder_path}' already exists.")

        # print(file_list)

        pdf_len = len(file_list)
        i = 0
        for i in range(pdf_len):
            pdf_path_folder = file_list[i]
            table_2_excel(pdf_path_folder, excel_path, output_path)



folder_path = 'C:\Ashutosh\Random\pdf2excel3'

pdf_path = "./sample2.pdf"

excel_path = "./path_temp/out2.xlsx"

output_path = "./result.xlsx"

# Type 1 to enter a file and 2 to enter a folder: 
file_or_folder = 2

if file_or_folder == 1:
    pdf_file_2_excel(pdf_path, excel_path, output_path)
elif file_or_folder == 2:
    file_list = folder_list(folder_path)
    pdf_folder_2_excel(folder_path, output_path, file_list)
    # print(file_list)
    # pdf_folder_2_excel(excel_path, output_path, file_list)
    # pdf_len = len(file_list)
    # i = 0
    # for i in range(pdf_len):
    #     pdf_path_folder = file_list[i]
    #     pdf_file_2_excel(pdf_path_folder, excel_path, output_path)
        # pdf_folder_2_excel(pdf_path_folder, excel_path, output_path, file_list)
else:
    print("Invalid input. Choose file or folder")



