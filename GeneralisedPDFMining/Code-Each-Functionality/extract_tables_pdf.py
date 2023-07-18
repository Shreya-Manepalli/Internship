import locale
import PySimpleGUI as sg
import tabula
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles.borders import Border, Side
import pandas as pd

sg.theme('Purple')
SYMBOL_UP = '▲'
SYMBOL_DOWN = '▼'

def collapse(sectionLayout, visible, key):
    return sg.pin(sg.Column(sectionLayout, visible=visible, key=key))

# ---------------MAIN LAYOUT---------------
layout = [
    [sg.Text('Thousand & Decimal separator in PDF document:', font=(None, 9, 'bold'))],
    [sg.Radio('1.234.567,999', 'RADIO1', default=True, size=(19, 1), key='locale'),
     sg.Radio('1,234,567.999', 'RADIO1', default=False, size=(19, 1))],

    [sg.Text('Source for PDF document:', size=(25, 1), font=(None, 9, 'bold')),
     sg.InputText(key='pdfDocument'), sg.FileBrowse()],

    [sg.Text('Source of Excel file:', size=(25, 1), font=(None, 9, 'bold'))],
    [sg.Radio('Create a blank Excel file', 'RADIO0', default=False, key='newWb')],

    [sg.Text('Choose any of the following options:', font=(None, 9, 'bold'))],
    [sg.Checkbox('Apply cell border (All Borders)', default=False, key='border')],
    [sg.Checkbox('Copy text format (Bold, Italic & Underline)', default=False, key='formatText')],
    [sg.Checkbox('Convert cell "-" to 0', default=False, key='hyphenToZero')],

    [sg.Text('Choose folder to save:', size=(25, 1), font=(None, 9, 'bold')),
     sg.InputText(key='savedFolder'), sg.FolderBrowse()],
    [sg.Text('Enter file name (Ex: file.xlsx):', size=(25, 1), font=(None, 9, 'bold')),
     sg.InputText(default_text='file.xlsx', key='savedName')],

    [sg.Button('Submit'), sg.Button('Exit')],
    [sg.Output(size=(80, 3))]
]


def pdfToExcel(dict):
    if dict['newWb']:
        wb = Workbook()
    tables = tabula.read_pdf(dict['pdfDocument'], pages='all', multiple_tables=True)

    if dict['locale']:
        locale.setlocale(locale.LC_NUMERIC, "en_DK.UTF-8")  
    else:
        locale.setlocale(locale.LC_NUMERIC, "en_US.UTF-8")  

    
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    
    for table_num, table in enumerate(tables):
        
        sheet = wb.create_sheet(title=f'Table {table_num + 1}')
        excel_writer = pd.ExcelWriter((dict['savedFolder'] + '/' + dict['savedName']), engine='xlsxwriter')

    for i, table in enumerate(tables):
    
        sheet_name = f"Table {i+1}"
        table.to_excel(excel_writer, sheet_name=sheet_name, index=False)

    excel_writer.save()

        
    for i, row in enumerate(tables):
            
            for j, cell in enumerate(row):
                
                excelCell = sheet.cell(row=i + 1, column=j + 1)

                
                try:
                    excelCell.value = locale.atof(cell.replace('(', '-').replace(')', ''))
                except ValueError:
                    if dict['hyphenToZero'] and cell == "-":
                        excelCell.value = 0
                    else:
                        excelCell.value = cell

                
                if dict['border']:
                    excelCell.border = thin_border

                
                if dict['formatText']:
                    excelCell.font = Font(bold=False,
                                          italic=False,
                                          underline=False)

    
    if not dict['savedName'].endswith('.xlsx'):
        dict['savedName'] += '.xlsx'
    wb.save((dict['savedFolder'] + '/' + dict['savedName']))


window = sg.Window('Copy PDF tables to Excel', layout)

while True:  
    event, values = window.read()
    if event == sg.WIN_CLOSED or event == 'Exit':
        break
    if event == 'Submit':
        
        print('Running............')

        
        keys = ['newWb', 'pdfDocument',
                'locale',
                'border', 'hyphenToZero', 'formatText',
                'savedFolder', 'savedName'
                ]
        convertDict = {x: window[x].get() for x in keys}

        
        if convertDict['pdfDocument'] and convertDict['savedFolder']:
            pdfToExcel(convertDict)

        
        print('Done!!!!')

    if event.startswith('-OPEN SEC1-'):
        opened1 = not opened1
        window['-OPEN SEC1-'].update(SYMBOL_UP if opened1 else SYMBOL_DOWN)
        window['-SEC1-'].update(visible=opened1)

window.close()